import os
import sqlite3
from openpyxl import load_workbook
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR , "db" , "../db/air.db")
DATA_DIR = os.path.join(BASE_DIR, "data")

CITY_FILTERS = [
    "wroc", "bydg", "toru", "lubl", "gorz", "ziel",
    "lodz", "krak", "wars", "opol", "rzes", "bial",
    "gdan", "kato", "kiel", "olsz", "pozn", "szcz"
]


def load_excel_file(filepath, pollutant):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = ws[2]  # drugi wiersz: nazwy stacji
    stations = [cell.value for cell in header_row][1:]  # pomijamy pierwszą kolumnę

    for row in ws.iter_rows(min_row=6, values_only=True):  # dane od wiersza 6
        timestamp = row[0]

        if not isinstance(timestamp, datetime):
            continue

        date_part = timestamp.date().isoformat()
        hour = timestamp.hour
        values = row[1:]

        for station, value in zip(stations, values):
            if station is None or value is None:
                continue

            station_l = station.lower()

            # filtracja: tylko miasta wojewódzkie
            if not any(code in station_l for code in CITY_FILTERS):
                continue

            try:
                val = float(value)
            except:
                continue

            cur.execute(
                """
                INSERT INTO measurements (station_code, date, hour, pollutant, value)
                VALUES (?, ?, ?, ?, ?)
                """,
                (station, date_part, hour, pollutant, val)
            )

    conn.commit()
    conn.close()


def load_all_excels():
    files = os.listdir(DATA_DIR)

    mapping = {
        "PM10": "PM10",
        "PM25": "PM25",
        "NO2": "NO2",
        "SO2": "SO2",
        "O3": "O3"
    }

    for filename in files:
        fname_upper = filename.upper()
        fpath = os.path.join(DATA_DIR, filename)

        for key in mapping:
            if key in fname_upper:
                print(f"Ładowanie {filename} jako {mapping[key]}")
                load_excel_file(fpath, mapping[key])


if __name__ == "__main__":
    load_all_excels()
